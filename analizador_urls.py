# Codigo elaborado por: Jairo Santana García
# Pablo de Jesus García Medina
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import requests
import os
from virus_total_apis import PublicApi
import time

# Inicializamos el archivo de excel
lib = Workbook()
hoja1 = lib.active
hoja1.title = "Analisis de Virus"
# Declaramos el nombre de algunas celdas
hoja1["A1"] = "URL"
hoja1["B1"] = "Fecha de analisis"
hoja1["C1"] = "Total de analisis"
hoja1["D1"] = "Analisis positivos"
hoja1["E1"] = "Clasificacion"
# Solicitamos la API
key = input("Ingrese su API KEY: ")
# Hacemos uso del modulo virus_total_apis
api = PublicApi(key)
# Declaramos la variable para poner los valores en la fila 2
filas = 2

# Solicitamos al usuario el tipo de API que tiene
print("Seleccione que tipo de API Key tiene: ")
api_type = int(input("""[1] API de paga (Busquedas ilimitadas)
[2] API Gratuita (Aceptamos máximo 12 links)
Opcion: """))

# Hacemos una busqueda en nuestros archivos para encontar un txt especifico
for dirpath, dirnames, filenames in os.walk("."):
    for name in filenames:
        if "urls_sospechosas.txt" in name:
            # Abrimos el txt y lo leemos linea por linea
            y = open(os.path.join(dirpath, name), "r")
            linea = [linea.rstrip() for linea in y]
            count = 1
            for lin in linea:
                if lin != '':
                    # Si no cuenta con API premium le damos unos valores
                    # Para que espere 1 min el programa y pueda leer 4
                    # solicitudes mas, ya que el limite es 4 por minuto
                    if (count == 5 or count == 9) and api_type == 2:
                        print("""Como se estan realizando mas de 4 solicitudes
esperaremos un minuto para poder continuar""")
                        time.sleep(60)
                    virus = api.get_url_report(lin)
                    print(virus)
                    # Obtenemos el json o diccionario y lo ponemos en las celdas
                    hoja1.cell(filas, 1, virus["results"]["url"])
                    hoja1.cell(filas, 2, virus["results"]["scan_date"])
                    hoja1.cell(filas, 3, virus["results"]["total"])
                    hoja1.cell(filas, 4, virus["results"]["positives"])
                    # Analizamos los virus potenciales que tiene
                    # Para darle una clasificacion
                    if (virus["results"]["positives"] >= 0 or
                            virus["results"]["positives"] <= 3):
                        hoja1.cell(filas, 5, "Baja")
                    elif (virus["results"]["positives"] > 3 or
                          virus["results"]["positives"] <= 10):
                        hoja1.cell(filas, 5, "Media")
                    elif virus["results"]["positives"] > 10:
                        hoja1.cell(filas, 5, "Alta")
                    # En caso de algun error que de negativos o bien
                    # Algun valor no numerico se pone error en la celda
                    else: 
                        hoja1.cell(filas, 5, "ERROR")
                    filas += 1
                    count += 1
            # Cerramos el txt
            y.close()

# Guardamos el archivo
lib.save("reporte_analizador_urls.xlsx")
